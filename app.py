import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from datetime import datetime

# Function to load data from uploaded CSV


def load_data(file):
    data = pd.read_csv(file)
    return data

# Function for EDA (Exploratory Data Analysis)


def perform_eda(data):
    st.subheader("Data Summary")
    st.write(data.describe())

    st.subheader("Missing Data Check")
    st.write(data.isnull().sum())

    st.subheader("Correlation Heatmap")
    corr = data.corr()
    fig = plt.figure(figsize=(10, 6))
    sns.heatmap(corr, annot=True, cmap='coolwarm')
    st.pyplot(fig)

    st.subheader("Distribution of Columns")
    numeric_columns = data.select_dtypes(include=['float64', 'int64']).columns
    for col in numeric_columns:
        fig = plt.figure(figsize=(8, 4))
        sns.histplot(data[col], kde=True)
        st.pyplot(fig)


# Function for Attendance Tracking (Check-in and Summary)
attendance_data = []


def mark_attendance(employee_id, status):
    global attendance_data
    date = datetime.now().strftime("%Y-%m-%d")
    attendance_data.append(
        {'EmployeeID': employee_id, 'Date': date, 'Status': status})
    st.success(f"Attendance marked for Employee {employee_id} on {date}")


def display_attendance():
    df_attendance = pd.DataFrame(attendance_data)
    st.subheader("Attendance Summary")
    st.write(df_attendance)

# Function for generating Excel Report with Summary and Visualization


def generate_report(filtered_data, summary, chart_path, output_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Report"

    sheet['A1'] = "Customized Employee Report"
    sheet['A1'].font = Font(size=14, bold=True)

    sheet['A3'] = "Summary:"
    row = 4
    for key, value in summary.items():
        sheet[f'A{row}'] = key
        sheet[f'B{row}'] = value
        row += 1

    sheet['A7'] = "Filtered Data:"
    for col_num, column_name in enumerate(filtered_data.columns, start=1):
        sheet.cell(row=8, column=col_num, value=column_name)

    for row_num, row_data in enumerate(filtered_data.values, start=9):
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    img = Image(chart_path)
    sheet.add_image(img, "E3")

    wb.save(output_path)
    st.success(f"Report saved as {output_path}")

# Streamlit Web Interface


def main():
    st.title("Customized Report Generator with EDA and Attendance Tracking")

    # File upload section
    uploaded_file = st.file_uploader("Upload a CSV file", type=["csv"])
    if uploaded_file is not None:
        data = load_data(uploaded_file)
        st.subheader("Data Preview")
        st.write(data.head())

        # EDA section
        if st.checkbox("Perform Exploratory Data Analysis"):
            perform_eda(data)

        # Filtering data based on user input
        st.subheader("Filter Data")
        department = st.text_input("Department (optional)")
        date = st.date_input("Date (optional)", value=datetime.today())

        if st.button("Apply Filter"):
            filtered_data = data
            if department:
                filtered_data = filtered_data[filtered_data['Department']
                                              == department]
            if date:
                filtered_data = filtered_data[filtered_data['Date'] == str(
                    date)]
            st.write(filtered_data)

            # Generating report and chart
            if st.button("Generate Report"):
                summary = {
                    "Total Employees": filtered_data['EmployeeID'].nunique(),
                    "Total Hours Worked": filtered_data['HoursWorked'].sum(),
                    "Average Performance": filtered_data['PerformanceScore'].mean()
                }
                chart_path = "hours_worked_chart.png"
                fig = plt.figure(figsize=(8, 4))
                sns.barplot(x=filtered_data['Name'],
                            y=filtered_data['HoursWorked'])
                plt.title("Hours Worked by Employees")
                plt.tight_layout()
                plt.savefig(chart_path)
                plt.close()

                generate_report(filtered_data, summary,
                                chart_path, "employee_report.xlsx")

        # Attendance Tracking section
        st.subheader("Attendance Tracking")
        employee_id = st.text_input("Employee ID for Attendance")
        status = st.radio("Attendance Status", ["Present", "Absent"])

        if st.button("Mark Attendance"):
            mark_attendance(employee_id, status)

        if st.checkbox("View Attendance Summary"):
            display_attendance()


if __name__ == "__main__":
    main()
